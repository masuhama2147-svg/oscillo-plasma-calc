"""Load oscilloscope waveforms from the Nomura-lab xlsx format.

The source file (`オシロスコープ測定結果.xlsx`) has 5 sheets:
  - `重ね合わせ` (overlay of voltage waveforms across pulse-width conditions)
  - `PW目盛{0.50,1.00,1.50,2.00}` (TIME / CH1=V / CH2=A, 10 000 rows)

This loader skips the overlay sheet and yields one Waveform per measurement sheet.
Pulse-width metadata is picked up from cells W16:Y17 when present.
"""
from __future__ import annotations

from pathlib import Path
import numpy as np
import openpyxl

from .schema import Waveform


MEASUREMENT_SHEET_PREFIX = "PW目盛"


def list_xlsx_sheets(xlsx_path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _extract_pulse_width(ws) -> float | None:
    for row in range(15, 20):
        for col_letter in ("W", "X", "Y"):
            val = ws[f"{col_letter}{row}"].value
            if isinstance(val, (int, float)) and 0.05 < float(val) < 5.0:
                return float(val)
    return None


def load_xlsx(xlsx_path: str | Path,
              sheet_name: str | None = None) -> list[Waveform]:
    """Return one Waveform per measurement sheet. If `sheet_name` is given,
    only that sheet is loaded."""
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        waveforms: list[Waveform] = []
        for ws in wb.worksheets:
            if sheet_name is not None and ws.title != sheet_name:
                continue
            if not ws.title.startswith(MEASUREMENT_SHEET_PREFIX):
                continue

            t_list: list[float] = []
            v_list: list[float] = []
            i_list: list[float] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    continue  # header
                if row is None or len(row) < 4:
                    continue
                try:
                    t = float(row[0]); v = float(row[2]); i = float(row[3])
                except (TypeError, ValueError):
                    continue
                t_list.append(t); v_list.append(v); i_list.append(i)

            pw = _extract_pulse_width(ws)
            meta = {"source_file": str(xlsx_path), "sheet": ws.title}
            if pw is not None:
                meta["pulse_width_us"] = pw

            waveforms.append(Waveform(
                t=np.array(t_list),
                v=np.array(v_list),
                i=np.array(i_list),
                label=ws.title,
                meta=meta,
            ))
        return waveforms
    finally:
        wb.close()
